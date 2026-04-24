"""
Build-time script: reads the consumer master Excel and writes a fast lookup
cache next to the data files:

    Masterdata/cache/dcs.json         # tiny — list of unique DCs + column headers
    Masterdata/cache/consumers.db     # SQLite — O(1) per-IVRS lookup

In production (Render free tier) the FastAPI app hits this cache instead of
parsing the 24 MB / 96k-row spreadsheet on every request. SQLite is used for
per-IVRS lookups because loading the full 89 MB JSON into memory is too slow
and too memory-heavy for a 512 MB container.

Run from repo root or from backend/:
    python -m scripts.build_master_cache
    python backend/scripts/build_master_cache.py
"""
from __future__ import annotations

import json
import os
import sqlite3
import sys
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
BACKEND_DIR = HERE.parent
MASTERDATA_DIR = BACKEND_DIR / "Masterdata"
CACHE_DIR = MASTERDATA_DIR / "cache"

IVRS_COL = 8  # Column H
DC_COL = 4    # Column D


def main() -> int:
    consumer_master_dir = MASTERDATA_DIR / "Consumer Master"
    xlsx_files = sorted(p for p in consumer_master_dir.iterdir() if p.suffix.lower() == ".xlsx")
    if not xlsx_files:
        print(f"No .xlsx file found in {consumer_master_dir}", file=sys.stderr)
        return 1
    xlsx = xlsx_files[0]
    print(f"Reading {xlsx.name} ...")

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter)
    headers = [str(c).strip() if c is not None else "" for c in header_row]
    print(f"  {len(headers)} columns in header row")

    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    dcs_file = CACHE_DIR / "dcs.json"
    db_file = CACHE_DIR / "consumers.db"

    if db_file.exists():
        db_file.unlink()
    conn = sqlite3.connect(str(db_file))
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE consumers (
            ivrs TEXT PRIMARY KEY,
            dc   TEXT,
            data TEXT NOT NULL
        )
        """
    )
    # Bulk-insert in one transaction; very fast.
    cur.execute("BEGIN")

    dcs: list[str] = []
    dc_seen: set[str] = set()
    total = 0
    batch: list[tuple[str, str, str]] = []

    for row in rows_iter:
        padded = list(row) + [None] * (len(headers) - len(row))
        ivrs_val = padded[IVRS_COL - 1] if IVRS_COL - 1 < len(padded) else None
        ivrs = str(ivrs_val).strip() if ivrs_val is not None else ""
        if not ivrs:
            continue

        obj: dict = {}
        for h, v in zip(headers, padded):
            if not h:
                continue
            if v is None:
                obj[h] = None
            elif isinstance(v, (int, float, bool)):
                obj[h] = v
            else:
                obj[h] = str(v)

        dc_val = padded[DC_COL - 1] if DC_COL - 1 < len(padded) else None
        dc = str(dc_val).strip() if dc_val is not None else ""

        batch.append((ivrs, dc, json.dumps(obj, separators=(",", ":"))))
        total += 1
        if len(batch) >= 5000:
            cur.executemany("INSERT OR REPLACE INTO consumers(ivrs, dc, data) VALUES (?, ?, ?)", batch)
            batch.clear()

        if dc and dc not in dc_seen:
            dc_seen.add(dc)
            dcs.append(dc)

    if batch:
        cur.executemany("INSERT OR REPLACE INTO consumers(ivrs, dc, data) VALUES (?, ?, ?)", batch)
    cur.execute("CREATE INDEX idx_consumers_dc ON consumers(dc)")
    conn.commit()
    conn.execute("VACUUM")
    conn.close()

    with dcs_file.open("w") as f:
        json.dump({"dcs": dcs, "headers": headers}, f)

    print(f"Unique DCs ({len(dcs)}): {dcs}")
    print(f"Total consumers indexed: {total}")
    for p in (dcs_file, db_file):
        size_mb = round(p.stat().st_size / 1024 / 1024, 2)
        print(f"  wrote {p.relative_to(BACKEND_DIR)}  ({size_mb} MB)")
    return 0


if __name__ == "__main__":
    os.chdir(BACKEND_DIR)
    raise SystemExit(main())
