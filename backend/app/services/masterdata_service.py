from __future__ import annotations

import json
import subprocess
import sys
from functools import lru_cache
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from app.core.config import MASTERDATA_DIR

CACHE_DIR = MASTERDATA_DIR / "cache"
DCS_CACHE_FILE = CACHE_DIR / "dcs.json"
CONSUMERS_CACHE_FILE = CACHE_DIR / "consumers.json"


def _ensure_cache_built() -> None:
    """
    If the JSON cache is missing (first run after a fresh checkout), build it
    on-demand so every subsequent call is instant. Safe to call repeatedly.
    """
    if DCS_CACHE_FILE.exists() and CONSUMERS_CACHE_FILE.exists():
        return
    script = Path(__file__).resolve().parents[2] / "scripts" / "build_master_cache.py"
    if not script.exists():
        return
    try:
        subprocess.run([sys.executable, str(script)], check=True)
    except Exception:
        # If cache build fails we fall back to streaming the Excel.
        pass


@lru_cache(maxsize=1)
def _load_dcs_cache() -> dict | None:
    _ensure_cache_built()
    if not DCS_CACHE_FILE.exists():
        return None
    try:
        with DCS_CACHE_FILE.open() as f:
            return json.load(f)
    except Exception:
        return None


@lru_cache(maxsize=1)
def _load_consumers_cache() -> dict | None:
    _ensure_cache_built()
    if not CONSUMERS_CACHE_FILE.exists():
        return None
    try:
        with CONSUMERS_CACHE_FILE.open() as f:
            return json.load(f)
    except Exception:
        return None


@lru_cache(maxsize=1)
def _consumer_master_path() -> Path | None:
    consumer_master_dir = MASTERDATA_DIR / "Consumer Master"
    if not consumer_master_dir.exists():
        return None
    xlsx_files = sorted([p for p in consumer_master_dir.iterdir() if p.suffix.lower() == ".xlsx"])
    return xlsx_files[0] if xlsx_files else None


def _list_single_file(folder: Path, exts: Iterable[str]) -> Path | None:
    if not folder.exists():
        return None
    exts_lower = {e.lower().lstrip(".") for e in exts}
    for p in sorted(folder.iterdir()):
        if not p.is_file():
            continue
        if p.suffix.lower().lstrip(".") in exts_lower:
            return p
    return None


def _sorted_seal_images(folder: Path) -> list[Path]:
    if not folder.exists():
        return []
    allowed_exts = {"heic", "jpg", "jpeg", "png", "webp"}
    paths = [p for p in folder.iterdir() if p.is_file() and p.suffix.lower().lstrip(".") in allowed_exts]
    # Ensure deterministic mapping. If names contain numbers, this naturally keeps order.
    return sorted(paths, key=lambda p: p.name.lower())


@lru_cache(maxsize=1)
def load_consumer_master_rows() -> list[dict[str, Any]]:
    """
    Reads the first .xlsx inside `backend/Masterdata/Consumer Master/`.
    Fixed mapping based on your note:
      - Column D => DC name
      - Column H => IVRS
    """
    xlsx = _consumer_master_path()
    if not xlsx:
        return []

    workbook = load_workbook(xlsx, read_only=True, data_only=True)
    sheet = workbook.active

    rows: list[dict[str, Any]] = []

    # openpyxl is 1-indexed. Column D is index 4. Column H is index 8.
    dc_col_idx = 4
    ivrs_col_idx = 8

    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        # Skip header row (row 1). We assume first row has column headers.
        if i == 1:
            continue

        dc_val = row[dc_col_idx - 1] if dc_col_idx - 1 < len(row) else None
        ivrs_val = row[ivrs_col_idx - 1] if ivrs_col_idx - 1 < len(row) else None

        dc = str(dc_val).strip() if dc_val is not None else ""
        ivrs = str(ivrs_val).strip() if ivrs_val is not None else ""
        if not ivrs:
            continue

        rows.append(
            {
                "dc_name": dc,
                "ivrs": ivrs,
            }
        )

    return rows


def load_consumer_master_slice(limit: int = 50, offset: int = 0) -> list[dict[str, Any]]:
    """
    Fast path for /assigned: streams only a slice of rows from Excel instead of reading the whole file.
    Offset/limit are in terms of *data rows* (excluding the header row).
    """
    xlsx = _consumer_master_path()
    if not xlsx:
        return []

    workbook = load_workbook(xlsx, read_only=True, data_only=True)
    sheet = workbook.active

    dc_col_idx = 4
    ivrs_col_idx = 8

    start = max(offset, 0)
    # hard cap to keep response manageable
    lim = max(min(limit, 200), 1)
    end = start + lim

    out: list[dict[str, Any]] = []
    data_row_idx = 0

    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if i == 1:
            continue  # header

        if data_row_idx < start:
            data_row_idx += 1
            continue
        if data_row_idx >= end:
            break

        dc_val = row[dc_col_idx - 1] if dc_col_idx - 1 < len(row) else None
        ivrs_val = row[ivrs_col_idx - 1] if ivrs_col_idx - 1 < len(row) else None

        dc = str(dc_val).strip() if dc_val is not None else ""
        ivrs = str(ivrs_val).strip() if ivrs_val is not None else ""
        data_row_idx += 1
        if not ivrs:
            continue

        out.append({"dc_name": dc, "ivrs": ivrs})

    return out


@lru_cache(maxsize=1)
def load_unique_dcs(limit: int = 500) -> list[str]:
    """
    Returns unique DC names from Column D.
    Prefers the pre-built JSON cache; falls back to streaming the Excel.
    """
    cached = _load_dcs_cache()
    if cached and isinstance(cached.get("dcs"), list):
        return list(cached["dcs"])[:limit]

    xlsx = _consumer_master_path()
    if not xlsx:
        return []

    workbook = load_workbook(xlsx, read_only=True, data_only=True)
    sheet = workbook.active

    dc_col_idx = 4
    seen: set[str] = set()
    out: list[str] = []

    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if i == 1:
            continue
        dc_val = row[dc_col_idx - 1] if dc_col_idx - 1 < len(row) else None
        dc = str(dc_val).strip() if dc_val is not None else ""
        if not dc or dc in seen:
            continue
        seen.add(dc)
        out.append(dc)
        if len(out) >= limit:
            break

    return out


@lru_cache(maxsize=1)
def load_headers() -> list[str]:
    """Return Excel header row (row 1), normalized to strings."""
    cached = _load_dcs_cache()
    if cached and isinstance(cached.get("headers"), list):
        return list(cached["headers"])

    xlsx = _consumer_master_path()
    if not xlsx:
        return []

    workbook = load_workbook(xlsx, read_only=True, data_only=True)
    sheet = workbook.active

    headers: list[str] = []
    first_row = next(sheet.iter_rows(values_only=True, max_row=1), None)
    if first_row is None:
        return []
    for cell in first_row:
        headers.append(str(cell).strip() if cell is not None else "")
    return headers


def get_full_consumer_row(ivrs: str) -> dict[str, Any] | None:
    """
    Look up a consumer row by IVRS. O(1) via JSON cache;
    falls back to a streaming Excel scan.
    """
    target = str(ivrs).strip()
    if not target:
        return None

    cached = _load_consumers_cache()
    if cached is not None:
        return cached.get(target)

    xlsx = _consumer_master_path()
    if not xlsx:
        return None

    headers = load_headers()
    workbook = load_workbook(xlsx, read_only=True, data_only=True)
    sheet = workbook.active

    ivrs_col_idx = 8

    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if i == 1:
            continue
        ivrs_val = row[ivrs_col_idx - 1] if ivrs_col_idx - 1 < len(row) else None
        row_ivrs = str(ivrs_val).strip() if ivrs_val is not None else ""
        if row_ivrs != target:
            continue

        out: dict[str, Any] = {}
        for j, val in enumerate(row):
            key = headers[j] if j < len(headers) and headers[j] else f"col_{j + 1}"
            if val is None:
                out[key] = None
            elif isinstance(val, (int, float, bool)):
                out[key] = val
            else:
                out[key] = str(val)
        return out

    return None


def ivrs_belongs_to_dc(ivrs: str, dc_name: str) -> bool:
    """
    Validates that the IVRS exists and its DC matches.
    O(1) via JSON cache; falls back to Excel streaming.
    """
    target_ivrs = str(ivrs).strip()
    target_dc = str(dc_name).strip()
    if not target_ivrs or not target_dc:
        return False

    row = get_full_consumer_row(target_ivrs)
    if row is not None:
        row_dc = row.get("DC")
        if isinstance(row_dc, str) and row_dc.strip() == target_dc:
            return True
        return False

    # Fallback (shouldn't normally run once cache is built).
    xlsx = _consumer_master_path()
    if not xlsx:
        return False

    workbook = load_workbook(xlsx, read_only=True, data_only=True)
    sheet = workbook.active

    dc_col_idx = 4
    ivrs_col_idx = 8

    for i, sheet_row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if i == 1:
            continue

        ivrs_val = sheet_row[ivrs_col_idx - 1] if ivrs_col_idx - 1 < len(sheet_row) else None
        row_ivrs = str(ivrs_val).strip() if ivrs_val is not None else ""
        if row_ivrs != target_ivrs:
            continue

        dc_val = sheet_row[dc_col_idx - 1] if dc_col_idx - 1 < len(sheet_row) else None
        row_dc = str(dc_val).strip() if dc_val is not None else ""
        return row_dc == target_dc

    return False


@lru_cache(maxsize=1)
def load_master_media() -> dict[str, Any]:
    """
    Uses the first matching media files under the Masterdata folders.
    If you later add per-work-order subfolders, we can extend this.
    """
    new_video = _list_single_file(MASTERDATA_DIR / "New Meter Video", exts={"mp4", "mov", "m4v"})
    old_video = _list_single_file(MASTERDATA_DIR / "Old Meter Video", exts={"mp4", "mov", "m4v"})

    # Your structure says seals are under "Seal Data". Your current filesystem listing shows HEIC images
    # in MASTERDATA_DIR root as well. We support both:
    seal_dir = MASTERDATA_DIR / "Seal Data"
    seals = _sorted_seal_images(seal_dir)
    if not seals:
        seals = _sorted_seal_images(MASTERDATA_DIR)

    return {
        "new_video": new_video,
        "old_video": old_video,
        "seals": seals,
    }


def get_work_order_from_master(work_order_id: str) -> dict[str, Any] | None:
    """
    `work_order_id` is treated as IVRS (your Column H).
    """
    # Stream scan to avoid loading the entire file into memory.
    xlsx = _consumer_master_path()
    if not xlsx:
        return None

    workbook = load_workbook(xlsx, read_only=True, data_only=True)
    sheet = workbook.active

    dc_col_idx = 4
    ivrs_col_idx = 8
    target = str(work_order_id).strip()

    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if i == 1:
            continue
        ivrs_val = row[ivrs_col_idx - 1] if ivrs_col_idx - 1 < len(row) else None
        ivrs = str(ivrs_val).strip() if ivrs_val is not None else ""
        if not ivrs or ivrs != target:
            continue
        dc_val = row[dc_col_idx - 1] if dc_col_idx - 1 < len(row) else None
        dc = str(dc_val).strip() if dc_val is not None else ""
        return {"dc_name": dc, "ivrs": ivrs}

    return None


def pick_seal_photos_by_fixed_order(seals: list[Path]) -> dict[str, Path | None]:
    """
    Workflow expects 4 seal photos:
      - meter_body_seal_photo
      - nic_seal_photo
      - terminal_seal_photo
      - box_seal_photo
    If you only have one batch of seals, we map in filename order.
    """
    keys = ["meter_body_seal_photo", "nic_seal_photo", "terminal_seal_photo", "box_seal_photo"]
    out: dict[str, Path | None] = {}
    for idx, key in enumerate(keys):
        out[key] = seals[idx] if idx < len(seals) else None
    return out

